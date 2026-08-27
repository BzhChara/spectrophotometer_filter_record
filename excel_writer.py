import csv
import os
import time

from openpyxl import Workbook
from openpyxl.styles import Font

from app_paths import get_app_dir


EXCEL_FONT_NAME = "微软雅黑"
TEMP_CSV_DIR_NAME = "_temp_csv"
STABLE_SHEET_NAME = "稳定强度"
CALIBRATED_SHEET_NAME = "校准强度"
RATIO_SHEET_NAME = "透过率"
ABSORBANCE_SHEET_NAME = "吸光度"
DETAIL_SHEET_NAME = "测量明细"
REPEATABILITY_LABELS = ["最小值", "最大值", "平均数", "样本标准差", "样本CV"]
POPULATION_LABELS = ["最小值", "最大值", "平均数", "总体标准差", "总体CV"]
DETAIL_HEADERS = [
    "时间",
    "波长(nm)",
    "通道",
    "次数",
    "空气基底",
    "空气平均值",
    "通道系数",
    "稳定强度",
    "校准强度",
    "透过率",
    "吸光度",
    "窗口帧数",
    "透过率样本标准差",
    "透过率样本CV",
    "透过率极差",
    "透过率斜率/帧",
    "透过率CV阈值",
    "透过率斜率阈值/帧",
    "透过率极差阈值",
    "空气范围容差",
    "判定",
]


def check_output_available(path: str):
    if not os.path.exists(path):
        return True

    temp_output = path + ".lockcheck"
    try:
        os.rename(path, temp_output)
        os.rename(temp_output, path)
    except OSError:
        print(f"输出文件被占用，请先关闭: {os.path.abspath(path)}")
        return False

    return True


def create_run_output_dir():
    run_date = time.strftime("%Y_%m_%d")
    run_time = time.strftime("%H%M%S")
    output_dir = os.path.join(get_app_dir(), "data", run_date, f"record_{run_time}")
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def resolve_output_path(output_dir: str, file_name: str):
    return os.path.join(output_dir, os.path.basename(file_name))


def raw_csv_path(raw_excel_path: str):
    temp_dir = os.path.join(os.path.dirname(raw_excel_path), TEMP_CSV_DIR_NAME)
    csv_name = f"{os.path.splitext(os.path.basename(raw_excel_path))[0]}.csv"
    return os.path.join(temp_dir, csv_name)


def _selected_indices(target_indices=None):
    return list(range(24)) if target_indices is None else list(target_indices)


def create_raw_csv(raw_excel_path: str, target_indices=None):
    selected_indices = _selected_indices(target_indices)
    csv_path = raw_csv_path(raw_excel_path)
    os.makedirs(os.path.dirname(csv_path), exist_ok=True)
    file = open(csv_path, "w", newline="", encoding="utf-8-sig")
    writer = csv.writer(file)
    writer.writerow(["时间"] + [f"CH{idx + 1}" for idx in selected_indices])
    file.flush()
    return file, writer, csv_path


def append_raw_csv_row(file, writer, values: list[float], target_indices=None):
    selected_indices = _selected_indices(target_indices)
    row = [time.strftime("%H:%M:%S")] + [f"{values[idx]:.6f}" for idx in selected_indices]
    writer.writerow(row)
    file.flush()


def create_raw_excel(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = ["时间"] + [f"CH{i}" for i in range(1, 25)]
    ws.append(headers)
    ws.freeze_panes = "B2"

    for cell in ws[1]:
        cell.font = Font(name=EXCEL_FONT_NAME, size=12, bold=True)

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, 26):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 12

    wb.save(path)
    return wb, ws


def raw_csv_to_excel(csv_path: str, excel_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.freeze_panes = "B2"

    with open(csv_path, "r", newline="", encoding="utf-8-sig") as file:
        reader = csv.reader(file)
        for row_index, row in enumerate(reader, start=1):
            if row_index == 1:
                ws.append(row)
            else:
                values = [row[0]] + [float(value) if value else None for value in row[1:]]
                ws.append(values)

            for cell in ws[row_index]:
                cell.font = Font(name=EXCEL_FONT_NAME, size=12, bold=row_index == 1)
            for cell in ws[row_index][1:]:
                cell.number_format = "0.000000"

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 12

    wb.save(excel_path)
    return wb, ws


def create_stable_excel(path: str, target_indices=None):
    wb = Workbook()
    ws = wb.active
    ws.title = STABLE_SHEET_NAME
    wb.create_sheet(CALIBRATED_SHEET_NAME)
    wb.create_sheet(RATIO_SHEET_NAME)
    wb.create_sheet(ABSORBANCE_SHEET_NAME)
    wb.create_sheet(DETAIL_SHEET_NAME)
    refresh_stable_excel(wb, ws, path, [[] for _ in range(24)], target_indices)
    return wb, ws


def _stable_history(stable_values: list, idx: int):
    value = stable_values[idx]
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _summary_stats(values: list[float], sample: bool):
    if not values:
        return [None, None, None, None, None]

    avg_value = sum(values) / len(values)
    if sample and len(values) < 2:
        std_value = None
    else:
        divisor = len(values) - 1 if sample else len(values)
        std_value = (sum((value - avg_value) ** 2 for value in values) / divisor) ** 0.5

    cv_value = std_value / abs(avg_value) if std_value is not None and avg_value else None
    return [min(values), max(values), avg_value, std_value, cv_value]


def repeatability_stats(values: list[float]):
    """同一通道多次测量的样本统计。"""
    return _summary_stats(values, sample=True)


def population_stats(values: list[float]):
    """本次所选全部通道的总体统计。"""
    return _summary_stats(values, sample=False)


def _write_cell(ws, row: int, column: int, value, bold: bool = False, number_format: str = "0.000000"):
    cell = ws.cell(row=row, column=column, value=value)
    cell.font = Font(name=EXCEL_FONT_NAME, size=12, bold=bold)
    if isinstance(value, (int, float)):
        cell.number_format = number_format
    return cell


def _reset_sheet(ws):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    if ws.max_column:
        ws.delete_cols(1, ws.max_column)


def _refresh_summary_sheet(ws, histories: dict[int, list[float]], selected_indices: list[int], value_format: str):
    max_count = max((len(values) for values in histories.values()), default=0)
    measurement_count = max(max_count, 1)
    stats_col = measurement_count + 3
    bottom_stats_start = len(selected_indices) + 3

    _reset_sheet(ws)

    _write_cell(ws, 1, 1, "通道", bold=True)
    for attempt in range(measurement_count):
        _write_cell(ws, 1, 2 + attempt, f"第{attempt + 1}次", bold=True)

    for offset, label in enumerate(REPEATABILITY_LABELS):
        _write_cell(ws, 1, stats_col + offset, label, bold=True)

    for row_idx, idx in enumerate(selected_indices, start=2):
        history = histories[idx]
        _write_cell(ws, row_idx, 1, f"CH{idx + 1}", bold=True)
        for attempt, value in enumerate(history, start=1):
            _write_cell(ws, row_idx, 1 + attempt, value, number_format=value_format)

        for offset, value in enumerate(repeatability_stats(history)):
            number_format = "0.00%" if offset == 4 else value_format
            _write_cell(ws, row_idx, stats_col + offset, value, number_format=number_format)

    for offset, label in enumerate(POPULATION_LABELS):
        _write_cell(ws, bottom_stats_start + offset, 1, label, bold=True)
        for attempt in range(measurement_count):
            values = [history[attempt] for history in histories.values() if len(history) > attempt]
            number_format = "0.00%" if offset == 4 else value_format
            _write_cell(
                ws,
                bottom_stats_start + offset,
                2 + attempt,
                population_stats(values)[offset],
                number_format=number_format,
            )

    ws.column_dimensions["A"].width = 12
    for col_idx in range(2, stats_col + len(REPEATABILITY_LABELS)):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 14

    ws.column_dimensions["A"].width = 12


def _ratio_histories(stable_records: list | None, selected_indices: list[int]):
    if stable_records is None:
        return {idx: [] for idx in selected_indices}
    return {
        idx: [record["ratio"] for record in _stable_history(stable_records, idx) if record.get("ratio") is not None]
        for idx in selected_indices
    }


def _calibrated_histories(stable_records: list | None, selected_indices: list[int]):
    if stable_records is None:
        return {idx: [] for idx in selected_indices}
    return {
        idx: [
            record["calibrated_value"]
            for record in _stable_history(stable_records, idx)
            if record.get("calibrated_value") is not None
        ]
        for idx in selected_indices
    }


def _absorbance_histories(stable_records: list | None, selected_indices: list[int]):
    if stable_records is None:
        return {idx: [] for idx in selected_indices}
    return {
        idx: [
            record["absorbance"]
            for record in _stable_history(stable_records, idx)
            if record.get("absorbance") is not None
        ]
        for idx in selected_indices
    }


def _refresh_detail_sheet(ws, stable_records: list | None, selected_indices: list[int]):
    _reset_sheet(ws)
    for column, header in enumerate(DETAIL_HEADERS, start=1):
        _write_cell(ws, 1, column, header, bold=True)

    rows = []
    if stable_records is not None:
        for idx in selected_indices:
            for record in _stable_history(stable_records, idx):
                rows.append((idx, record))

    rows.sort(key=lambda item: (item[1].get("timestamp") or "", item[0], item[1].get("attempt") or 0))
    for row, (idx, record) in enumerate(rows, start=2):
        values = [
            record.get("timestamp"),
            record.get("wavelength"),
            f"CH{idx + 1}",
            record.get("attempt"),
            record.get("baseline"),
            record.get("air_mean"),
            record.get("calibration_factor"),
            record.get("stable_value"),
            record.get("calibrated_value"),
            record.get("ratio"),
            record.get("absorbance"),
            record.get("sample_count"),
            record.get("ratio_sample_std"),
            record.get("ratio_cv"),
            record.get("ratio_range"),
            record.get("ratio_slope_per_frame"),
            record.get("ratio_cv_limit"),
            record.get("ratio_slope_limit"),
            record.get("ratio_range_limit"),
            record.get("air_tolerance"),
            record.get("result"),
        ]
        for column, value in enumerate(values, start=1):
            number_format = "0.000000"
            if column in {2, 4, 12}:
                number_format = "0"
            elif column in {10, 13, 14, 15, 16, 17, 18, 19, 20}:
                number_format = "0.0000%"
            _write_cell(ws, row, column, value, number_format=number_format)

    ws.freeze_panes = "A2"
    widths = [24, 12, 10, 10, 14, 14, 14, 14, 14, 14, 14, 12, 20, 18, 16, 20, 18, 22, 20, 16, 12]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=column).column_letter].width = width


def refresh_stable_excel(
    wb: Workbook,
    ws,
    path: str,
    stable_values: list,
    target_indices=None,
    stable_records: list | None = None,
):
    selected_indices = _selected_indices(target_indices)
    histories = {idx: _stable_history(stable_values, idx) for idx in selected_indices}
    calibrated_histories = _calibrated_histories(stable_records, selected_indices)
    ratio_histories = _ratio_histories(stable_records, selected_indices)
    absorbance_histories = _absorbance_histories(stable_records, selected_indices)

    _refresh_summary_sheet(ws, histories, selected_indices, "0.000000")

    calibrated_ws = (
        wb[CALIBRATED_SHEET_NAME]
        if CALIBRATED_SHEET_NAME in wb.sheetnames
        else wb.create_sheet(CALIBRATED_SHEET_NAME)
    )
    _refresh_summary_sheet(calibrated_ws, calibrated_histories, selected_indices, "0.000000")

    ratio_ws = wb[RATIO_SHEET_NAME] if RATIO_SHEET_NAME in wb.sheetnames else wb.create_sheet(RATIO_SHEET_NAME)
    _refresh_summary_sheet(ratio_ws, ratio_histories, selected_indices, "0.0000%")

    absorbance_ws = (
        wb[ABSORBANCE_SHEET_NAME]
        if ABSORBANCE_SHEET_NAME in wb.sheetnames
        else wb.create_sheet(ABSORBANCE_SHEET_NAME, 3)
    )
    _refresh_summary_sheet(absorbance_ws, absorbance_histories, selected_indices, "0.000000")

    detail_ws = wb[DETAIL_SHEET_NAME] if DETAIL_SHEET_NAME in wb.sheetnames else wb.create_sheet(DETAIL_SHEET_NAME)
    _refresh_detail_sheet(detail_ws, stable_records, selected_indices)

    wb.save(path)
