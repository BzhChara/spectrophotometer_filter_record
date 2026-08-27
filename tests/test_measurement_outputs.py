import math
import os
import struct
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from config_loader import DEFAULT_OPTIONS, load_config
from excel_writer import (
    ABSORBANCE_SHEET_NAME,
    CALIBRATED_SHEET_NAME,
    DETAIL_SHEET_NAME,
    RATIO_SHEET_NAME,
    create_stable_excel,
    population_stats,
    refresh_stable_excel,
    repeatability_stats,
)
from filter_logic import (
    absorbance_from_transmittance,
    air_window_is_stable,
    calibrate_intensity,
    channel_gain_calibration,
    filter_is_inserted,
    filter_window_is_stable,
    window_diagnostics,
)


class StatisticsTests(unittest.TestCase):
    def test_repeatability_uses_sample_standard_deviation(self):
        minimum, maximum, average, standard_deviation, cv = repeatability_stats([1.0, 2.0, 3.0])

        self.assertEqual(minimum, 1.0)
        self.assertEqual(maximum, 3.0)
        self.assertEqual(average, 2.0)
        self.assertEqual(standard_deviation, 1.0)
        self.assertEqual(cv, 0.5)

    def test_single_repeatability_measurement_has_no_standard_deviation_or_cv(self):
        self.assertEqual(repeatability_stats([2.0]), [2.0, 2.0, 2.0, None, None])

    def test_channel_summary_uses_population_standard_deviation(self):
        values = population_stats([1.0, 2.0, 3.0])

        self.assertAlmostEqual(values[3], math.sqrt(2 / 3))
        self.assertAlmostEqual(values[4], math.sqrt(2 / 3) / 2)


class ChannelCalibrationTests(unittest.TestCase):
    def test_channel_gain_uses_selected_air_baseline_mean(self):
        baselines = [10.0, 20.0, 100.0]

        air_mean, factor = channel_gain_calibration(baselines, [0, 1], 0)

        self.assertEqual(air_mean, 15.0)
        self.assertEqual(factor, 1.5)
        self.assertEqual(calibrate_intensity(3.0, factor), 4.5)

    def test_channel_gain_requires_every_selected_baseline(self):
        self.assertIsNone(channel_gain_calibration([10.0, None], [0, 1], 0))

    def test_absorbance_uses_negative_base_10_logarithm_of_transmittance(self):
        self.assertAlmostEqual(absorbance_from_transmittance(0.3), -math.log10(0.3))
        self.assertIsNone(absorbance_from_transmittance(0.0))


class SerialCalibrationIntegrationTests(unittest.TestCase):
    def test_channel_gain_is_frozen_when_filter_insertion_is_detected(self):
        from spectro_ui.serial_worker import SerialTestWorker, TestConfig

        air_values = [10.0] * 24
        air_values[1] = 20.0
        filter_values = air_values.copy()
        filter_values[0] = 3.0
        filter_values[1] = 20.8
        frames = [
            self._frame(air_values),
            self._frame(air_values),
            self._frame(filter_values),
            self._frame(filter_values),
            self._frame(filter_values),
        ]

        config = TestConfig(
            port="COM_TEST",
            wavelength=520,
            channel_group=1,
            test_mode="filter",
            output="raw.xlsx",
            stable_output="stable.xlsx",
            filter_stable_window_frames=3,
            filter_stable_cv_limit=1.0,
            filter_stable_slope_limit=1.0,
            filter_stable_range_limit=1.0,
            air_tolerance=10.0,
            air_warmup_seconds=0.0,
            air_stable_window_frames=2,
            air_stable_cv_limit=1.0,
            air_stable_slope_limit=1.0,
            air_stable_range_limit=1.0,
            no_start=True,
            keep_light=True,
        )
        worker = SerialTestWorker(config)
        baseline_updates = []
        state_updates = []
        stable_updates = []
        event_log = []
        worker.baseline_value_changed.connect(
            lambda channel_index, value: (
                baseline_updates.append((channel_index, value)),
                event_log.append(("baseline", channel_index, value)),
            )
        )
        worker.channel_state_changed.connect(
            lambda channel_index, state: (
                state_updates.append((channel_index, state)),
                event_log.append(("state", channel_index, state)),
            )
        )
        worker.stable_value_changed.connect(
            lambda channel_index, stable, calibrated, ratio_percent, absorbance: stable_updates.append(
                (channel_index, stable, calibrated, ratio_percent, absorbance)
            )
        )

        fake_serial = self._FakeSerial(frames, worker)
        with tempfile.TemporaryDirectory() as temp_dir:
            with (
                patch("spectro_ui.serial_worker.serial.Serial", return_value=fake_serial),
                patch(
                    "spectro_ui.serial_worker.create_run_output_dir",
                    return_value=temp_dir,
                ),
            ):
                worker.run()

            saved = load_workbook(os.path.join(temp_dir, "stable.xlsx"), data_only=True)
            detail_ws = saved[DETAIL_SHEET_NAME]
            self.assertAlmostEqual(detail_ws["F2"].value, 130 / 12, delta=1e-5)
            self.assertAlmostEqual(detail_ws["G2"].value, 13 / 12, delta=1e-5)
            self.assertAlmostEqual(detail_ws["I2"].value, 3.25, delta=1e-5)
            self.assertAlmostEqual(detail_ws["K2"].value, -math.log10(0.3), delta=1e-5)
            saved.close()

        self.assertEqual(len(stable_updates), 1)
        self.assertEqual(stable_updates[0][0], 0)
        self.assertAlmostEqual(stable_updates[0][4], -math.log10(0.3), delta=1e-5)

        self.assertTrue(
            any(
                channel_index == 1 and abs(value - 20.8) <= 1e-5
                for channel_index, value in baseline_updates
            )
        )
        channel_one_states = [state for channel_index, state in state_updates if channel_index == 0]
        self.assertIn("ready", channel_one_states)
        self.assertIn("detecting", channel_one_states)
        self.assertIn("waiting_air", channel_one_states)
        self.assertLess(
            channel_one_states.index("ready"),
            channel_one_states.index("detecting"),
        )
        self.assertLess(
            channel_one_states.index("detecting"),
            channel_one_states.index("waiting_air"),
        )
        baseline_update_index = next(
            index
            for index, event in enumerate(event_log)
            if event[0] == "baseline"
            and event[1] == 1
            and abs(event[2] - 20.8) <= 1e-5
        )
        waiting_air_index = next(
            index
            for index, event in enumerate(event_log)
            if event == ("state", 0, "waiting_air")
        )
        self.assertLess(baseline_update_index, waiting_air_index)

    @staticmethod
    def _frame(values):
        payload = struct.pack("<24f", *values)
        frame_without_checksum = bytes([0x5A, 0xA5, 0x01, 0x56, len(payload)]) + payload
        return frame_without_checksum + bytes([sum(frame_without_checksum) & 0xFF])

    class _FakeSerial:
        def __init__(self, frames, worker):
            self._frames = list(frames)
            self._worker = worker
            self.is_open = True

        def read(self, _size):
            if self._frames:
                return self._frames.pop(0)
            self._worker.stop()
            return b""

        def reset_input_buffer(self):
            pass

        def write(self, data):
            return len(data)

        def close(self):
            self.is_open = False


class WindowDiagnosticsTests(unittest.TestCase):
    def test_filter_insertion_uses_the_air_range_lower_limit(self):
        self.assertTrue(filter_is_inserted(94.9, baseline=100.0, air_tolerance=0.05))
        self.assertFalse(filter_is_inserted(95.0, baseline=100.0, air_tolerance=0.05))
        self.assertFalse(filter_is_inserted(105.0, baseline=100.0, air_tolerance=0.05))

    def test_diagnostics_are_normalized_to_the_baseline(self):
        diagnostics = window_diagnostics([2.0, 3.0, 4.0], baseline=10.0)

        self.assertEqual(diagnostics["sample_count"], 3)
        self.assertAlmostEqual(diagnostics["ratio_sample_std"], 0.1)
        self.assertAlmostEqual(diagnostics["ratio_cv"], 1 / 3)
        self.assertAlmostEqual(diagnostics["ratio_range"], 0.2)
        self.assertAlmostEqual(diagnostics["ratio_slope_per_frame"], 0.1)
        self.assertNotIn("raw_range", diagnostics)

    def test_constant_air_window_is_stable_without_an_existing_baseline(self):
        self.assertTrue(
            air_window_is_stable(
                [100.0] * 20,
                baseline=None,
                cv_limit=0.0025,
                slope_limit=0.0003,
                range_limit=0.005,
                required_count=20,
            )
        )

    def test_air_window_rejects_excessive_slope(self):
        values = [100.0 + frame * 0.025 for frame in range(20)]

        self.assertFalse(
            air_window_is_stable(
                values,
                baseline=100.0,
                cv_limit=0.0025,
                slope_limit=0.0002,
                range_limit=0.01,
                required_count=20,
            )
        )

    def test_air_window_rejects_excessive_range(self):
        values = [99.7, 100.3] * 10

        self.assertFalse(
            air_window_is_stable(
                values,
                baseline=100.0,
                cv_limit=0.01,
                slope_limit=0.01,
                range_limit=0.005,
                required_count=20,
            )
        )

    def test_constant_filter_window_is_stable(self):
        self.assertTrue(
            filter_window_is_stable(
                [30.0] * 10,
                baseline=100.0,
                cv_limit=0.0025,
                slope_limit=0.0002,
                range_limit=0.0025,
                required_count=10,
            )
        )

    def test_filter_window_rejects_excessive_slope(self):
        values = [30.0 + frame * 0.025 for frame in range(10)]

        self.assertFalse(
            filter_window_is_stable(
                values,
                baseline=100.0,
                cv_limit=0.01,
                slope_limit=0.0002,
                range_limit=0.01,
                required_count=10,
            )
        )

    def test_filter_window_rejects_excessive_range(self):
        values = [29.85, 30.15] * 5

        self.assertFalse(
            filter_window_is_stable(
                values,
                baseline=100.0,
                cv_limit=0.01,
                slope_limit=0.01,
                range_limit=0.0025,
                required_count=10,
            )
        )


class RuntimeSmokeTests(unittest.TestCase):
    def test_serial_worker_imports_with_system_dependencies(self):
        from spectro_ui.serial_worker import SerialTestWorker, TestConfig

        self.assertIsNotNone(SerialTestWorker)
        filter_option_names = {
            "filter_stable_window_frames",
            "filter_stable_cv_limit",
            "filter_stable_slope_limit",
            "filter_stable_range_limit",
        }
        air_option_names = {
            "air_tolerance",
            "air_warmup_seconds",
            "air_stable_window_frames",
            "air_stable_cv_limit",
            "air_stable_slope_limit",
            "air_stable_range_limit",
        }
        self.assertTrue(filter_option_names.issubset(DEFAULT_OPTIONS))
        self.assertTrue(filter_option_names.issubset(TestConfig.__dataclass_fields__))
        self.assertTrue(air_option_names.issubset(DEFAULT_OPTIONS))
        self.assertTrue(air_option_names.issubset(TestConfig.__dataclass_fields__))
        self.assertEqual(DEFAULT_OPTIONS["port"], "")
        self.assertEqual(DEFAULT_OPTIONS["test_mode"], "filter")
        self.assertIn("test_mode", TestConfig.__dataclass_fields__)
        self.assertNotIn("filter_ratio", DEFAULT_OPTIONS)
        self.assertNotIn("ratio_tolerance", DEFAULT_OPTIONS)
        self.assertNotIn("filter_ratio", TestConfig.__dataclass_fields__)
        self.assertNotIn("ratio_tolerance", TestConfig.__dataclass_fields__)
        self.assertEqual(DEFAULT_OPTIONS["filter_stable_window_frames"], 10)
        self.assertEqual(DEFAULT_OPTIONS["filter_stable_cv_limit"], 0.25)
        self.assertEqual(DEFAULT_OPTIONS["filter_stable_slope_limit"], 0.02)
        self.assertEqual(DEFAULT_OPTIONS["filter_stable_range_limit"], 0.25)
        self.assertEqual(DEFAULT_OPTIONS["air_stable_window_frames"], 20)

    def test_legacy_filter_ratio_is_migrated_to_test_mode(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "config.ini")
            with open(path, "w", encoding="utf-8") as file:
                file.write("[settings]\nfilter_ratio = 0\nratio_tolerance = 5\n")

            values = load_config(path)

        self.assertEqual(values["test_mode"], "air")
        self.assertNotIn("filter_ratio", values)
        self.assertNotIn("ratio_tolerance", values)

    def test_project_config_exposes_adjustable_stability_options(self):
        project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        values = load_config(os.path.join(project_dir, "config.ini"))
        adjustable_names = {
            "test_mode",
            "filter_stable_window_frames",
            "filter_stable_cv_limit",
            "filter_stable_slope_limit",
            "filter_stable_range_limit",
            "air_tolerance",
            "air_warmup_seconds",
            "air_stable_window_frames",
            "air_stable_cv_limit",
            "air_stable_slope_limit",
            "air_stable_range_limit",
        }

        self.assertTrue(adjustable_names.issubset(values))
        self.assertEqual(values["port"], "")
        self.assertEqual(values["test_mode"], "filter")


class StableWorkbookTests(unittest.TestCase):
    def test_workbook_saves_stable_ratio_and_traceable_details(self):
        stable_values = [[3.0, 4.0], [6.0, 8.0]] + [[] for _ in range(22)]
        stable_records = [
            [
                self._record(attempt=1, baseline=10.0, stable_value=3.0),
                self._record(attempt=2, baseline=10.0, stable_value=4.0),
            ],
            [
                self._record(attempt=1, baseline=20.0, stable_value=6.0),
                self._record(attempt=2, baseline=20.0, stable_value=8.0),
            ],
        ] + [[] for _ in range(22)]

        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "stable.xlsx")
            wb, ws = create_stable_excel(path, [0, 1])
            refresh_stable_excel(wb, ws, path, stable_values, [0, 1], stable_records)

            saved = load_workbook(path, data_only=True)
            self.assertEqual(
                saved.sheetnames,
                [
                    "稳定强度",
                    CALIBRATED_SHEET_NAME,
                    RATIO_SHEET_NAME,
                    ABSORBANCE_SHEET_NAME,
                    DETAIL_SHEET_NAME,
                ],
            )

            stable_ws = saved["稳定强度"]
            self.assertAlmostEqual(stable_ws["H2"].value, math.sqrt(0.5))
            self.assertAlmostEqual(stable_ws["B8"].value, 1.5)

            calibrated_ws = saved[CALIBRATED_SHEET_NAME]
            self.assertAlmostEqual(calibrated_ws["B2"].value, 4.5)
            self.assertAlmostEqual(calibrated_ws["C2"].value, 6.0)
            self.assertAlmostEqual(calibrated_ws["B3"].value, 4.5)
            self.assertAlmostEqual(calibrated_ws["C3"].value, 6.0)

            ratio_ws = saved[RATIO_SHEET_NAME]
            self.assertAlmostEqual(ratio_ws["B2"].value, 0.3)
            self.assertAlmostEqual(ratio_ws["C3"].value, 0.4)

            absorbance_ws = saved[ABSORBANCE_SHEET_NAME]
            self.assertAlmostEqual(absorbance_ws["B2"].value, -math.log10(0.3))
            self.assertAlmostEqual(absorbance_ws["C3"].value, -math.log10(0.4))

            detail_ws = saved[DETAIL_SHEET_NAME]
            self.assertEqual(detail_ws.max_column, 21)
            self.assertNotIn("原始强度极差", [cell.value for cell in detail_ws[1]])
            self.assertNotIn("目标透过率", [cell.value for cell in detail_ws[1]])
            self.assertNotIn("比例容差", [cell.value for cell in detail_ws[1]])
            self.assertEqual(detail_ws["T1"].value, "空气范围容差")
            self.assertEqual(detail_ws["C2"].value, "CH1")
            self.assertEqual(detail_ws["E2"].value, 10.0)
            self.assertEqual(detail_ws["F1"].value, "空气平均值")
            self.assertEqual(detail_ws["F2"].value, 15.0)
            self.assertEqual(detail_ws["G1"].value, "通道系数")
            self.assertEqual(detail_ws["G2"].value, 1.5)
            self.assertEqual(detail_ws["H2"].value, 3.0)
            self.assertEqual(detail_ws["I1"].value, "校准强度")
            self.assertEqual(detail_ws["I2"].value, 4.5)
            self.assertAlmostEqual(detail_ws["J2"].value, 0.3)
            self.assertEqual(detail_ws["K1"].value, "吸光度")
            self.assertAlmostEqual(detail_ws["K2"].value, -math.log10(0.3))
            self.assertEqual(detail_ws["L1"].value, "窗口帧数")
            self.assertEqual(detail_ws["Q1"].value, "透过率CV阈值")
            self.assertEqual(detail_ws["R1"].value, "透过率斜率阈值/帧")
            self.assertEqual(detail_ws["S1"].value, "透过率极差阈值")
            self.assertAlmostEqual(detail_ws["Q2"].value, 0.0025)
            self.assertAlmostEqual(detail_ws["R2"].value, 0.0002)
            self.assertAlmostEqual(detail_ws["S2"].value, 0.0025)
            self.assertEqual(detail_ws["U2"].value, "通过")

    @staticmethod
    def _record(attempt: int, baseline: float, stable_value: float):
        air_mean = 15.0
        calibration_factor = air_mean / baseline
        return {
            "timestamp": "2026-08-27 12:34:56.789",
            "wavelength": 520,
            "attempt": attempt,
            "baseline": baseline,
            "air_mean": air_mean,
            "calibration_factor": calibration_factor,
            "stable_value": stable_value,
            "calibrated_value": stable_value * calibration_factor,
            "ratio": stable_value / baseline,
            "absorbance": -math.log10(stable_value / baseline),
            "sample_count": 10,
            "ratio_sample_std": 0.001,
            "ratio_cv": 0.003,
            "ratio_range": 0.002,
            "ratio_slope_per_frame": 0.0001,
            "ratio_cv_limit": 0.0025,
            "ratio_slope_limit": 0.0002,
            "ratio_range_limit": 0.0025,
            "air_tolerance": 0.05,
            "result": "通过",
        }


if __name__ == "__main__":
    unittest.main()
